import pandas as pd
import openpyxl
from os.path import exists

from scrape import Scrape
from analyze import Analyze


df = pd.read_excel('Input.xlsx')
listId = df['URL_ID'].tolist()
listUrl =  df['URL'].tolist()

def saveToExcel(index,data):
    print(index)
    srcfile = openpyxl.load_workbook('Output Data Structure.xlsx', read_only=False, keep_vba=True)
    sheetname = srcfile['Sheet1']
    sheetname.cell(row=index, column=3).value = data['positiveScore']
    sheetname.cell(row=index, column=4).value = data['negativeScore']
    sheetname.cell(row=index, column=5).value = data['polarityScore']
    sheetname.cell(row=index, column=6).value = data['subjectiveScore']
    sheetname.cell(row=index, column=7).value = data['avgSentenceLength']
    sheetname.cell(row=index, column=8).value = data['%OfComplexWords']
    sheetname.cell(row=index, column=9).value = data['fogIndex']
    sheetname.cell(row=index, column=10).value= data['avgNoOfWoordsPerSentence']
    sheetname.cell(row=index, column=11).value= data['countofComplexWords']
    sheetname.cell(row=index, column=12).value= data['wordCount']
    sheetname.cell(row=index, column=13).value= data['totalSyllable']
    sheetname.cell(row=index, column=14).value= data['personalPronounsCount']
    sheetname.cell(row=index, column=15).value = data['avgWordLength']
    #save the file
    srcfile.save('Output Data Structure.xlsx')
    

for index,element in enumerate(listUrl):
    if not exists(f'textFiles/{int(listId[index])}.txt'):
        Scrape.scrapeData(int(listId[index]),listUrl[index])
    

    data = (Analyze().returnAllAnalysedData(int(listId[index])))
    saveToExcel(int(listId[index]) + 1, data)


    